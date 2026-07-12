from __future__ import annotations

import argparse
import json
import re
import shutil
import unicodedata
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path.home() / "Downloads" / "Календарь тендеров.xlsx"
DATA_TARGET = ROOT / "tenderflow-admin" / "src" / "data" / "tender-calendar.ts"
PUBLIC_TARGET = ROOT / "tenderflow-admin" / "public" / "tender-calendar" / "qazcloud-tender-calendar-2026.xlsx"
SHAREPOINT_SOURCE = "https://freedomholdingcorporation.sharepoint.com/:x:/s/FreedomCloud464/IQDUOK5HQYjgTrZ4c_WZG37DAbUhHCOVJUhodAwfGxwNK-Y?e=8YQfyi"

URL_RE = re.compile(r"https?://[^\s;,)]+", re.IGNORECASE)
HYPERLINK_RE = re.compile(
    r'^=(?:HYPERLINK|ГИПЕРССЫЛКА)\(\s*"([^"]+)"\s*[,;]\s*"([^"]*)"\s*\)$',
    re.IGNORECASE,
)


@dataclass(frozen=True)
class CellValue:
    raw: Any
    text: str
    url: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).replace("\xa0", " ").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", " · ", text)
    return text.strip(" ·")


def first_nonempty(*values: Any) -> str:
    for value in values:
        text = clean_text(value)
        if text and text.lower() not in {"отсутствует", "нет данных", "nan", "none"}:
            return text
    return ""


def extract_url(value: Any) -> str:
    text = clean_text(value)
    match = URL_RE.search(text)
    return match.group(0) if match else ""


def parse_hyperlink_formula(value: Any) -> tuple[str, str]:
    if not isinstance(value, str):
        return "", ""
    match = HYPERLINK_RE.match(value.strip())
    if not match:
        return "", ""
    return match.group(2).replace('""', '"').strip(), match.group(1).strip()


def parse_date(value: Any, epoch: datetime | None = None) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 1 <= float(value) <= 100000:
        try:
            converted = from_excel(value, epoch=epoch) if epoch else from_excel(value)
            if isinstance(converted, time):
                return ""
            return converted.date().isoformat() if isinstance(converted, datetime) else converted.isoformat()
        except (TypeError, ValueError, OverflowError):
            return ""

    text = clean_text(value)
    if not text or "отсутств" in text.lower():
        return ""
    for pattern in (
        r"\b(\d{1,2})[./-](\d{1,2})[./-](\d{4})\b",
        r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b",
    ):
        match = re.search(pattern, text)
        if not match:
            continue
        try:
            if pattern.startswith("\\b(\\d{4})"):
                year, month, day = map(int, match.groups())
            else:
                day, month, year = map(int, match.groups())
            return date(year, month, day).isoformat()
        except ValueError:
            return ""
    return ""


def parse_money(value: Any) -> int | float | None:
    if value in (None, "", "-") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value != value:
            return None
        rounded = round(float(value), 2)
        return int(rounded) if rounded.is_integer() else rounded
    text = clean_text(value).lower()
    if not text or any(marker in text for marker in ("отсутств", "нет данных", "%")):
        return None
    compact = re.sub(r"[^0-9,.-]", "", text.replace(" ", ""))
    if not compact or compact in {"-", ".", ","}:
        return None
    if compact.count(",") == 1 and compact.count(".") == 0:
        compact = compact.replace(",", ".")
    else:
        compact = compact.replace(",", "")
    try:
        number = round(float(compact), 2)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def parse_percent(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 6)
    text = clean_text(value).replace("%", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return round(number / 100 if abs(number) > 1 else number, 6)


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean_text(value)).lower().replace("ё", "е")
    replacements = {
        "акционерное общество": "ао",
        "товарищество с ограниченной ответственностью": "тоо",
        "республиканское государственное учреждение": "ргу",
        "коммунальное государственное учреждение": "кгу",
        "некоммерческое акционерное общество": "нао",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[^a-zа-я0-9]+", " ", text).strip()


def join_unique(*values: Any, separator: str = " · ") -> str:
    items: list[str] = []
    seen: set[str] = set()
    for value in values:
        for part in clean_text(value).split(separator):
            item = part.strip()
            key = normalize(item)
            if item and key not in seen:
                items.append(item)
                seen.add(key)
    return separator.join(items)


def classify_customer(platform: str, holding: str) -> str:
    combined = normalize(f"{platform} {holding}")
    if "самрук" in combined:
        return "Самрук-Казына"
    if "государ" in combined or "goszakup" in combined:
        return "Государственный сектор"
    return "Прочие"


def is_yes(value: Any) -> bool:
    return normalize(value) in {"да", "yes", "1", "true", "топ", "топ 20"}


def canonical_status(value: Any, winner: Any = "") -> str:
    text = normalize(value)
    winner_text = normalize(winner)
    if "qazcloud" in winner_text or "казклауд" in winner_text:
        return "Выиграли (QazCloud)"
    if not text:
        return "Не указан"
    if any(marker in text for marker in ("не уча", "не подходим", "не подош", "не соответств", "нерентабель", "не рентабель", "проигр")):
        return "Не участвуем"
    if any(marker in text for marker in ("не состоя", "отмен")):
        return "Не состоялся" if "не состоя" in text else "Отменен"
    if any(marker in text for marker in ("выиг", "побед")):
        return "Завершен"
    if any(marker in text for marker in ("подались", "участвуем", "учавствуем", "подготовить подач")):
        return "Участвуем"
    if "прием заяв" in text:
        return "Прием заявок"
    if any(marker in text for marker in ("не опублик", "ожидаем", "ждем публика", "планируется", "декабрь 2026 тендер")):
        return "Планируется"
    if "опублик" in text:
        return "Опубликован"
    if any(marker in text for marker in ("заверш", "итоги")):
        return "Завершен"
    if any(marker in text for marker in ("работ", "расчет", "расчит", "просчит", "проработ", "подготов", "иб")):
        return "В работе"
    return "Не указан"


def row_template(source_sheet: str) -> dict[str, Any]:
    return {
        "id": "",
        "notion": "",
        "customer": "",
        "customerType": "Прочие",
        "holding": "",
        "city": "",
        "title": "",
        "service": "",
        "canCover": "",
        "initialAmount": None,
        "winnerAmount": None,
        "dumping": None,
        "top20": False,
        "winner": "",
        "qazcloud": "",
        "manager": "",
        "status": "Не указан",
        "tenderDate": "",
        "publishedAt": "",
        "deadlineAt": "",
        "openingAt": "",
        "source": source_sheet,
        "sourceSheet": source_sheet,
        "contractNumber": "",
        "contractAmount": None,
        "contractEnd": "",
        "nextTenderDate": "",
        "purchaseMethod": "",
        "notes": "",
        "tenderNumber": "",
        "bidSecurity": None,
        "contractSecurity": None,
        "contactName": "",
        "contactInfo": "",
        "sourceLink": "",
    }


class SheetReader:
    def __init__(self, ws_formula: Any, ws_values: Any, epoch: datetime):
        self.ws_formula = ws_formula
        self.ws_values = ws_values
        self.epoch = epoch
        self.headers: dict[str, int] = {}
        for cell in ws_formula[1]:
            key = clean_text(cell.value)
            if key and key not in self.headers:
                self.headers[key] = cell.column

    def cell(self, row: int, header: str) -> CellValue:
        column = self.headers.get(header)
        if not column:
            return CellValue(None, "", "")
        formula_cell = self.ws_formula.cell(row=row, column=column)
        value_cell = self.ws_values.cell(row=row, column=column)
        friendly, formula_url = parse_hyperlink_formula(formula_cell.value)
        hyperlink_url = formula_cell.hyperlink.target if formula_cell.hyperlink else ""
        raw = value_cell.value if value_cell.value is not None else formula_cell.value
        text = first_nonempty(friendly, value_cell.value, formula_cell.value)
        if friendly:
            text = friendly
        return CellValue(raw, text, first_nonempty(hyperlink_url, formula_url, extract_url(text)))

    def text(self, row: int, header: str) -> str:
        return self.cell(row, header).text

    def url(self, row: int, header: str) -> str:
        return self.cell(row, header).url

    def date(self, row: int, header: str) -> str:
        cell = self.cell(row, header)
        return parse_date(cell.raw, self.epoch) or parse_date(cell.text, self.epoch)

    def money(self, row: int, header: str) -> int | float | None:
        cell = self.cell(row, header)
        return parse_money(cell.raw) if parse_money(cell.raw) is not None else parse_money(cell.text)


def add_common_fields(row: dict[str, Any]) -> dict[str, Any]:
    row["customer"] = clean_text(row["customer"])
    row["title"] = clean_text(row["title"])
    row["manager"] = first_nonempty(row["manager"], "Не назначен")
    row["status"] = first_nonempty(row["status"], "Не указан")
    row["tenderDate"] = first_nonempty(row["openingAt"], row["deadlineAt"], row["publishedAt"])
    row["customerType"] = classify_customer(row["source"], row["holding"])
    row["qazcloud"] = "Да" if "qazcloud" in normalize(row["winner"]) or "казклауд" in normalize(row["winner"]) else ""
    return row


def parse_main(reader: SheetReader) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index in range(2, reader.ws_formula.max_row + 1):
        row = row_template("Календарь тендеров IaaS")
        row.update(
            customer=reader.text(index, "Заказчик"),
            title=reader.text(index, "Наименование тендера"),
            holding=reader.text(index, "Холдинг или группа"),
            service=first_nonempty(reader.text(index, "Тип услуги"), reader.text(index, "Услуга")),
            initialAmount=reader.money(index, "Выделенная сумма тендера"),
            winnerAmount=reader.money(index, "Сумма "),
            top20=is_yes(reader.text(index, "ТОП")),
            winner=reader.text(index, "Победитель "),
            manager=reader.text(index, "Ответственный"),
            status=canonical_status(
                reader.text(index, "Статус участия\n(Подались/ Идет рассмотрение)"),
                reader.text(index, "Победитель "),
            ),
            publishedAt=reader.date(index, "Публикование"),
            deadlineAt=reader.date(index, "День окончания приема заявок"),
            openingAt=reader.date(index, "Дата"),
            purchaseMethod=reader.text(index, "Площадка закупки"),
            tenderNumber=reader.text(index, "Номер тендера"),
            bidSecurity=reader.money(index, "Обеспечение заявки"),
            contractSecurity=reader.money(index, "Обеспечение договора"),
            contactName=reader.text(index, "Контактное лицо"),
            contactInfo=reader.text(index, "Контакты"),
            notion=first_nonempty(reader.url(index, "Ссылка на Notion"), reader.text(index, "Ссылка на Notion")),
            sourceLink=first_nonempty(
                reader.url(index, "Наименование тендера"),
                reader.url(index, "Площадка закупки"),
            ),
            notes=join_unique(reader.text(index, "Примечания"), reader.text(index, "Статус проработки")),
            dumping=parse_percent(reader.cell(index, "Процент снижения").raw),
        )
        add_common_fields(row)
        if row["customer"] and row["title"]:
            rows.append(row)
    return rows


def parse_top(reader: SheetReader) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    audit: list[dict[str, Any]] = []
    for index in range(2, reader.ws_formula.max_row + 1):
        title = reader.text(index, "Наименование тендера")
        customer = reader.text(index, "Заказчик")
        if not customer or not title:
            continue
        published_at = reader.date(index, "Дата")
        deadline_at = reader.date(index, "Сроки")
        row = row_template("ТОП тендеров")
        row.update(
            customer=customer,
            title=title,
            initialAmount=reader.money(index, "Выделенная сумма тендера"),
            winnerAmount=reader.money(index, "Сумма "),
            top20=True,
            winner=reader.text(index, "Победитель "),
            manager=first_nonempty(reader.text(index, "Менеджер тендерного"), reader.text(index, "Хантер")),
            status=canonical_status(
                first_nonempty(reader.text(index, "Статус тендерного отдела"), reader.text(index, "Статус менеджера")),
                reader.text(index, "Победитель "),
            ),
            publishedAt=published_at,
            deadlineAt=deadline_at,
            purchaseMethod=reader.text(index, "Площадка закупки"),
            notes=join_unique(
                reader.text(index, "Статус тендерного отдела"),
                reader.text(index, "Статус менеджера"),
                reader.text(index, "Задача"),
                reader.text(index, "План закупок"),
            ),
            sourceLink=first_nonempty(
                reader.url(index, "Наименование тендера"),
                reader.url(index, "Площадка закупки"),
                reader.url(index, "План закупок"),
            ),
        )
        add_common_fields(row)
        rows.append(row)
        audit.append(
            {
                "announcement": f"TOP-{len(audit) + 1:03d}",
                "lot": row["tenderNumber"],
                "publishedAt": published_at,
                "deadlineAt": deadline_at,
                "method": row["purchaseMethod"],
                "title": title,
                "description": row["notes"],
                "amount": row["initialAmount"] or row["winnerAmount"],
                "region": row["city"],
                "organizer": customer,
                "status": row["status"],
                "url": row["sourceLink"],
                "mergeResult": "Слито с общим календарём; флаг Топ-20 сохранён",
            }
        )
    return rows, audit


def parse_ib_2025(reader: SheetReader) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index in range(2, reader.ws_formula.max_row + 1):
        row = row_template("ИБ 2025")
        row.update(
            customer=reader.text(index, "Организатор"),
            title=reader.text(index, "Название"),
            city=reader.text(index, "Регион"),
            service="Информационная безопасность",
            initialAmount=reader.money(index, "Стоимость"),
            winnerAmount=reader.money(index, "Сумма"),
            winner=reader.text(index, "Победитель"),
            status=canonical_status(reader.text(index, "Статус"), reader.text(index, "Победитель")),
            publishedAt=reader.date(index, "Дата"),
            purchaseMethod=reader.text(index, "Источник"),
            sourceLink=first_nonempty(reader.url(index, "Название"), reader.url(index, "Источник")),
            notes=join_unique(
                f"БИН организатора: {reader.text(index, 'БИН Организатора')}" if reader.text(index, "БИН Организатора") else "",
                f"Публикация в 2026: {reader.text(index, 'Публикование в 2026 году')}" if reader.text(index, "Публикование в 2026 году") else "",
            ),
            dumping=parse_percent(reader.cell(index, "Процент снижения").raw),
        )
        add_common_fields(row)
        if row["customer"] and row["title"]:
            rows.append(row)
    return rows


def parse_ib(reader: SheetReader) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index in range(2, reader.ws_formula.max_row + 1):
        row = row_template("ИБ")
        row.update(
            customer=reader.text(index, "Заказчик"),
            title=reader.text(index, "Наименование тендера"),
            service="Информационная безопасность",
            initialAmount=reader.money(index, "Выделенная сумма тендера"),
            status=canonical_status(reader.text(index, "Статус участия\n(Подались/ Идет рассмотрение)")),
            publishedAt=reader.date(index, "Дата"),
            purchaseMethod=reader.text(index, "Площадка закупки"),
            bidSecurity=reader.money(index, "Обеспечение заявки"),
            contractSecurity=reader.money(index, "Обеспечение договора"),
            sourceLink=first_nonempty(reader.url(index, "Наименование тендера"), reader.url(index, "Площадка закупки")),
        )
        add_common_fields(row)
        if row["customer"] and row["title"]:
            rows.append(row)
    return rows


def parse_almaty(reader: SheetReader) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index in range(2, reader.ws_formula.max_row + 1):
        unit_price = reader.money(index, "Цена за единицу")
        row = row_template("Объем Алматы")
        row.update(
            customer=reader.text(index, "Организатор"),
            title=reader.text(index, "Название"),
            city=reader.text(index, "Регион"),
            initialAmount=reader.money(index, "Стоимость"),
            status=canonical_status(reader.text(index, "Статус")),
            publishedAt=reader.date(index, "Начало"),
            deadlineAt=reader.date(index, "Завершение"),
            purchaseMethod=reader.text(index, "Способ проведения"),
            sourceLink=first_nonempty(reader.url(index, "Название"), reader.url(index, "Источник")),
            notes=join_unique(
                f"БИН организатора: {reader.text(index, 'БИН Организатора')}" if reader.text(index, "БИН Организатора") else "",
                f"Цена за единицу: {unit_price}" if unit_price is not None else "",
                f"Место поставки: {reader.text(index, 'Место поставки')}" if reader.text(index, "Место поставки") else "",
            ),
        )
        add_common_fields(row)
        if row["customer"] and row["title"]:
            rows.append(row)
    return rows


def parse_existing_data(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8")
    prefix = "export const tenderCalendarData = "
    start = text.find(prefix)
    end = text.rfind(" as const;")
    if start < 0 or end < 0:
        return {}
    return json.loads(text[start + len(prefix) : end])


def stable_tender_id(row: dict[str, Any]) -> str:
    link = clean_text(row.get("sourceLink"))
    for pattern in (
        r"tenderplus\.kz/zakupki/(\d+)",
        r"announce/index/(\d+)",
        r"publics/lot/(\d+)",
        r"popup:item/(\d+)",
    ):
        match = re.search(pattern, link, re.IGNORECASE)
        if match:
            return f"{pattern.split('\\\\')[0]}:{match.group(1)}"
    return ""


def date_distance(left: dict[str, Any], right: dict[str, Any]) -> int | None:
    left_value = first_nonempty(left.get("deadlineAt"), left.get("publishedAt"), left.get("openingAt"))
    right_value = first_nonempty(right.get("deadlineAt"), right.get("publishedAt"), right.get("openingAt"))
    if not left_value or not right_value:
        return None
    try:
        return abs((date.fromisoformat(left_value) - date.fromisoformat(right_value)).days)
    except ValueError:
        return None


def is_duplicate(left: dict[str, Any], right: dict[str, Any]) -> bool:
    left_id = stable_tender_id(left)
    right_id = stable_tender_id(right)
    if left_id and right_id and left_id == right_id:
        return True
    if normalize(left.get("customer")) != normalize(right.get("customer")):
        return False
    if normalize(left.get("title")) != normalize(right.get("title")):
        return False
    distance = date_distance(left, right)
    if distance is not None:
        return distance <= 90
    return left.get("sourceSheet") != right.get("sourceSheet")


def choose_date(current: str, incoming: str, mode: str) -> str:
    if not current:
        return incoming
    if not incoming:
        return current
    if mode == "earliest":
        return min(current, incoming)
    return max(current, incoming)


def merge_rows(target: dict[str, Any], incoming: dict[str, Any]) -> None:
    for key, value in incoming.items():
        if key in {"id", "source", "sourceSheet", "notes", "top20", "publishedAt", "deadlineAt", "openingAt", "initialAmount", "winnerAmount"}:
            continue
        if target.get(key) in (None, "", "Не указан", "Не назначен") and value not in (None, ""):
            target[key] = value
    target["top20"] = bool(target.get("top20") or incoming.get("top20"))
    target["source"] = join_unique(target.get("source"), incoming.get("source"), separator=" + ")
    target["sourceSheet"] = join_unique(target.get("sourceSheet"), incoming.get("sourceSheet"), separator=" + ")
    target["notes"] = join_unique(target.get("notes"), incoming.get("notes"))
    target["publishedAt"] = choose_date(target.get("publishedAt", ""), incoming.get("publishedAt", ""), "earliest")
    target["deadlineAt"] = choose_date(target.get("deadlineAt", ""), incoming.get("deadlineAt", ""), "latest")
    target["openingAt"] = choose_date(target.get("openingAt", ""), incoming.get("openingAt", ""), "latest")
    for key in ("initialAmount", "winnerAmount"):
        values = [value for value in (target.get(key), incoming.get(key)) if isinstance(value, (int, float))]
        target[key] = max(values) if values else None
    add_common_fields(target)


def deduplicate(rows: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    result: list[dict[str, Any]] = []
    buckets: dict[tuple[str, str], list[int]] = {}
    link_index: dict[str, int] = {}
    removed = 0

    for source_row in rows:
        row = deepcopy(source_row)
        link_id = stable_tender_id(row)
        duplicate_index = link_index.get(link_id) if link_id else None
        key = (normalize(row.get("customer")), normalize(row.get("title")))
        if duplicate_index is None:
            for candidate_index in buckets.get(key, []):
                if is_duplicate(result[candidate_index], row):
                    duplicate_index = candidate_index
                    break
        if duplicate_index is None:
            duplicate_index = len(result)
            result.append(row)
            buckets.setdefault(key, []).append(duplicate_index)
            if link_id:
                link_index[link_id] = duplicate_index
        else:
            merge_rows(result[duplicate_index], row)
            removed += 1
            if link_id:
                link_index[link_id] = duplicate_index
    return result, removed


def nearest_deadline_key(row: dict[str, Any]) -> tuple[int, int, str, str]:
    today = date.today()
    value = row.get("deadlineAt")
    if value:
        try:
            parsed = date.fromisoformat(value)
            if parsed >= today:
                return (0, (parsed - today).days, normalize(row.get("customer")), normalize(row.get("title")))
            return (1, (today - parsed).days, normalize(row.get("customer")), normalize(row.get("title")))
        except ValueError:
            pass
    fallback = first_nonempty(row.get("publishedAt"), row.get("openingAt"), row.get("tenderDate"))
    try:
        parsed = date.fromisoformat(fallback)
        return (2, -parsed.toordinal(), normalize(row.get("customer")), normalize(row.get("title")))
    except (TypeError, ValueError):
        return (3, 0, normalize(row.get("customer")), normalize(row.get("title")))


def in_h2_2026(value: str) -> bool:
    return bool(value and "2026-07-01" <= value <= "2026-12-31")


def in_2026(value: str) -> bool:
    return bool(value and value.startswith("2026-"))


def validate_calendar(data: dict[str, Any]) -> None:
    rows = data["calendar"]
    date_fields = ("tenderDate", "publishedAt", "deadlineAt", "openingAt", "contractEnd", "nextTenderDate")
    money_fields = ("initialAmount", "winnerAmount", "contractAmount", "bidSecurity", "contractSecurity")
    iso_date = re.compile(r"^\d{4}-\d{2}-\d{2}$")

    for index, row in enumerate(rows, start=1):
        if not row["customer"] or not row["title"]:
            raise ValueError(f"Empty calendar row at position {index}")
        if not isinstance(row["status"], str):
            raise TypeError(f"Status must be text at position {index}")
        for field in date_fields:
            value = row[field]
            if value and (not isinstance(value, str) or not iso_date.fullmatch(value)):
                raise ValueError(f"Invalid {field} at position {index}: {value!r}")
        for field in money_fields:
            value = row[field]
            if value is not None and (isinstance(value, bool) or not isinstance(value, (int, float))):
                raise TypeError(f"Invalid {field} at position {index}: {value!r}")

    duplicate_check, removed = deduplicate(rows)
    if removed or len(duplicate_check) != len(rows):
        raise ValueError(f"Duplicate rows remain after import: {removed}")
    if rows != sorted(rows, key=nearest_deadline_key):
        raise ValueError("Calendar rows are not sorted by the nearest submission deadline")


def import_calendar(source: Path) -> dict[str, Any]:
    existing = parse_existing_data(DATA_TARGET)
    formula_book = load_workbook(source, data_only=False, read_only=False)
    value_book = load_workbook(source, data_only=True, read_only=False)

    readers = {
        name: SheetReader(formula_book[name], value_book[name], formula_book.epoch)
        for name in formula_book.sheetnames
    }

    main_rows = parse_main(readers["Календарь тендеров IaaS"])
    top_rows, top_audit = parse_top(readers["ТОП тендеров"])
    ib_2025_rows = parse_ib_2025(readers["ИБ 2025"])
    ib_rows = parse_ib(readers["ИБ"])
    almaty_rows = parse_almaty(readers["Объем Алматы"])
    excel_rows = main_rows + top_rows + ib_2025_rows + ib_rows + almaty_rows

    preserved_renewals: list[dict[str, Any]] = []
    for old_row in existing.get("calendar", []):
        if old_row.get("contractEnd") or old_row.get("nextTenderDate") or "договоры qazcloud" in normalize(old_row.get("source")):
            normalized_row = row_template(first_nonempty(old_row.get("sourceSheet"), old_row.get("source"), "Договоры QazCloud 2026"))
            normalized_row.update(old_row)
            normalized_row.setdefault("publishedAt", "")
            normalized_row.setdefault("deadlineAt", "")
            normalized_row.setdefault("openingAt", old_row.get("tenderDate", ""))
            normalized_row.setdefault("sourceLink", "")
            add_common_fields(normalized_row)
            if normalized_row["customer"] and normalized_row["title"]:
                preserved_renewals.append(normalized_row)

    merged_rows, removed = deduplicate(excel_rows + preserved_renewals)
    while True:
        merged_rows, additional_removed = deduplicate(merged_rows)
        if additional_removed == 0:
            break
        removed += additional_removed
    merged_rows.sort(key=nearest_deadline_key)
    for index, row in enumerate(merged_rows, start=1):
        row["id"] = f"T-{index:03d}"

    samruk_contracts = existing.get("samrukContracts", [])
    for contract in samruk_contracts:
        if not contract.get("documentUrl"):
            contract["documentUrl"] = SHAREPOINT_SOURCE

    metrics = {
        "source_rows_excel": len(excel_rows),
        "source_rows_main": len(main_rows),
        "source_rows_top": len(top_rows),
        "source_rows_ib_2025": len(ib_2025_rows),
        "source_rows_ib": len(ib_rows),
        "source_rows_almaty": len(almaty_rows),
        "preserved_renewals": len(preserved_renewals),
        "original_count": len(excel_rows) + len(preserved_renewals),
        "final_count": len(merged_rows),
        "deduped_removed": removed,
        "duplicate_count": 0,
        "top20_count": sum(1 for row in merged_rows if row["top20"]),
        "published_dates": sum(1 for row in merged_rows if row["publishedAt"]),
        "deadline_dates": sum(1 for row in merged_rows if row["deadlineAt"]),
        "opening_dates": sum(1 for row in merged_rows if row["openingAt"]),
        "amount_rows": sum(1 for row in merged_rows if row["initialAmount"] is not None or row["winnerAmount"] is not None),
        "h2_2026": sum(
            1
            for row in merged_rows
            if any(in_h2_2026(row[field]) for field in ("publishedAt", "deadlineAt", "openingAt", "nextTenderDate"))
        ),
        "renewals_2026": sum(1 for row in merged_rows if in_2026(row["contractEnd"]) or in_2026(row["nextTenderDate"])),
        "samruk_contracts": len(samruk_contracts),
    }
    control = [
        {"label": "Строк из Excel до очистки", "value": metrics["source_rows_excel"]},
        {"label": "Строк после объединения", "value": metrics["final_count"]},
        {"label": "Дублей удалено", "value": metrics["deduped_removed"]},
        {"label": "Пустых строк в реестре", "value": 0},
        {"label": "Строк с датой публикации", "value": metrics["published_dates"]},
        {"label": "Строк со сроком подачи", "value": metrics["deadline_dates"]},
        {"label": "Строк с датой вскрытия / итогов", "value": metrics["opening_dates"]},
        {"label": "Строк с числовой суммой", "value": metrics["amount_rows"]},
        {"label": "Строк с флагом Топ-20", "value": metrics["top20_count"]},
        {"label": "Тендеров во втором полугодии 2026", "value": metrics["h2_2026"]},
        {"label": "Будущих перезаключений в 2026", "value": metrics["renewals_2026"]},
        {"label": "Договоров Самрук-Казына", "value": metrics["samruk_contracts"]},
    ]

    return {
        "generatedAt": datetime.now().replace(microsecond=0).isoformat(),
        "xlsxPath": "/tender-calendar/qazcloud-tender-calendar-2026.xlsx",
        "metrics": metrics,
        "calendar": merged_rows,
        "samrukContracts": samruk_contracts,
        "top20Audit": top_audit,
        "control": control,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import the tender calendar workbook into the admin portal.")
    parser.add_argument("source", nargs="?", type=Path, default=DEFAULT_SOURCE)
    args = parser.parse_args()
    source = args.source.expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(source)

    data = import_calendar(source)
    validate_calendar(data)
    DATA_TARGET.write_text(
        "/* Auto-generated from Календарь тендеров.xlsx. Do not hand-edit row data. */\n"
        "export const tenderCalendarData = "
        + json.dumps(data, ensure_ascii=False, indent=2)
        + " as const;\n",
        encoding="utf-8",
    )
    PUBLIC_TARGET.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, PUBLIC_TARGET)
    print(json.dumps(data["metrics"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
